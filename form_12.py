from tkinter import *
from tkinter import ttk
import time

root = Tk()
root.title('Начало работы')
ttk.Label(root, text='\n  Программа запустится в фоновом режиме.  \n\n\
  Закройте это окно, чтобы начать расчет.  \n\n\
  Средняя скорость обработки - 2200 строк формы 074 в секунду.  \n', font=("Arial", 13)).grid()
root.mainloop()

start_time = time.clock()

'''
ABSTRACTION
'''
class Patient:
    def __init__(self, visit_type, MKB10):
        self.visit_type = str(visit_type)
        self.MKB10 = str(MKB10)

'''
IMPORT PATIENTS INFO
'''
import os
import openpyxl

patients = []
files = []

for filename in os.listdir(os.getcwd()):
    f = filename.split('.')
    if f[-1] == 'xlsx' and f[0] != 'results':
        files.append(filename)

for filename in files:
    main_wb = openpyxl.load_workbook(filename)
    sheet = main_wb['Sheet1']
    max_patient = sheet.max_row

    for i in range(max_patient-2):
        patient = Patient(
            sheet.cell(row=i+3, column=3).value,
            sheet.cell(row=i+3, column=10).value)
        
        if patient.visit_type == '1':
            patients.append(patient)

'''
TABLE LINES
'''
#A00-T98
l2 = []
for i in range(65, 65+20):
    for j in range(100):
        if j > 9:
            l2.append(chr(i) + str(j))
        else:
            l2.append(chr(i) + '0' + str(j))

#A00-B99
l3 = []
for i in range(65, 67):
    for j in range(100):
        if j > 9:
            l3.append(chr(i) + str(j))
        else:
            l3.append(chr(i) + '0' + str(j))

l4 = ['B18.0.1']

l5 = ['B18.2']

#C00-D48
l6 = []
for i in range(100):
    if i > 9:
        l6.append('C' + str(i))
    else:
        l6.append('C' + '0' + str(i))
for i in range(49):
    if i > 9:
        l6.append('D' + str(i))
    else:
        l6.append('D' + '0' + str(i))

l7 = ['D22', 'D23', 'D28.0', 'D29.0', 'D29.2', 'D29.4']

l8 = ['D27']

#D50-D89
l9 = ['D'+str(i) for i in range(50, 90)]

#D50-D64
l10 = ['D'+str(i) for i in range(50, 65)]

l11 = ['D50']

l12 = ['D60', 'D61']

l13 = ['D66', 'D67', 'D68.0.1.4']

#D80-D84
l14 = ['D'+str(i) for i in range(80, 85)]

l15 = ['D86']

l16 = ['D89']

#E00-E90
l17 = []
for i in range(91):
    if i > 9:
        l17.append('E' + str(i))
    else:
        l17.append('E' + '0' + str(i))

l18 = ['E01.0', 'E04.0', 'E04.1']

l19 = ['E01.0', 'E04.0']

l20 = ['E01.8', 'E03']

l21 = ['E01.1.2', 'E04.1.2']

l22 = ['E05']

l23 = ['E06']

#E10-E14
l24 = ['E'+str(i) for i in range(10, 15)]

l25 = ['E23.2']

l26 = ['E66']

l27 = ['E89.0']

#F00-F99
l28 = []
for i in range(100):
    if i > 9:
        l28.append('F' + str(i))
    else:
        l28.append('F' + '0' + str(i))

#G00-G99
l29 = []
for i in range(100):
    if i > 9:
        l29.append('G' + str(i))
    else:
        l29.append('G' + '0' + str(i))

l30 = ['G00', 'G03', 'G04', 'G06', 'G08', 'G09']

l31 = ['G40', 'G41']

l32 = ['G50', 'G51', 'G52', 'G54', 'G56', 'G57', 'G58', \
       'G60', 'G61', 'G62', 'G64', 'G70', 'G71', 'G72']

l33 = ['G80']

#H00-H59
l34 = []
for i in range(60):
    if i > 9:
        l34.append('H' + str(i))
    else:
        l34.append('H' + '0' + str(i))

l35 = ['H10', 'H11']

l36 = ['H25', 'H26']

l37 = ['H49', 'H50']

l38 = ['H52.1']

#H60-H95
l39 = ['H'+str(i) for i in range(60, 96)]

l40 = ['H65', 'H66', 'H68', 'H69', 'H70', 'H71', 'H72', 'H73', 'H74']

l41 = ['H65.0.1', 'H66.0']

l42 = ['H65.2.3.4', 'H66.1.2.3']

#I00-I99
l43 = []
for i in range(100):
    if i > 9:
        l43.append('I' + str(i))
    else:
        l43.append('I' + '0' + str(i))

l44 = ['I00', 'I01', 'I02']

l45 = ['I00']

l46 = ['I05', 'I06', 'I07', 'I08', 'I09']

l47 = ['I05', 'I06', 'I07', 'I08']

l48 = ['I10']

l49 = ['I34', 'I35', 'I36', 'I37', 'I38']

#J00-J99
l50 = []
for i in range(100):
    if i > 9:
        l50.append('J' + str(i))
    else:
        l50.append('J' + '0' + str(i))

l51 = ['J02', 'J03']

l52 = ['J04']

l53 = ['J12', 'J13', 'J14', 'J15', 'J16', 'J18']

l54 = ['J30.1', 'J30.2', 'J30.3', 'J30.4']

l55 = ['J31']

l56 = ['J35']

l57 = ['J37']

l58 = ['J41', 'J42']

l59 = ['J45', 'J46']

#K00-K93
l60 = []
for i in range(94):
    if i > 9:
        l60.append('K' + str(i))
    else:
        l60.append('K' + '0' + str(i))

l61 = ['K21']

l62 = ['K25', 'K26', 'K27']

l63 = ['K29']

l64 = ['K30']

l65 = ['K31']

l66 = ['K50']

l67 = ['K51']

l68 = ['K58']

l69 = ['K73', 'K75.2', 'K75.3']

l70 = ['K80']

l71 = ['K81', 'K83.0']

l72 = ['K85', 'K86']

l73 = ['K90.0']

#L00-L99
l74 = []
for i in range(100):
    if i > 9:
        l74.append('L' + str(i))
    else:
        l74.append('L' + '0' + str(i))

#L00-L08
l75 = ['L0' + str(i) for i in range(9)]

l76 = ['L20']

l77 = ['L23', 'L24', 'L25']

#M00-M99
l78 = []
for i in range(100):
    if i > 9:
        l78.append('M' + str(i))
    else:
        l78.append('M' + '0' + str(i))

l79 = ['M05', 'M06', 'M08.0']

l80 = ['M32']

#N00-N99
l81 = []
for i in range(100):
    if i > 9:
        l81.append('N' + str(i))
    else:
        l81.append('N' + '0' + str(i))

l82 = ['N00']

l83 = ['N03']

l84 = ['N10', 'N11', 'N12']

l85 = ['N11']

l86 = ['N30']

l87 = ['N30.0']

l88 = ['N30.1.2']

l89 = ['N43']

l90 = ['N91', 'N92', 'N94']

#O00-O99 except O80
l91 = []
for i in range(100):
    if i > 9:
        l91.append('O' + str(i))
    else:
        l91.append('O' + '0' + str(i))
l91.remove('O80')

#P05-P96
l92 = []
for i in range(5, 97):
    if i > 9:
        l92.append('O' + str(i))
    else:
        l92.append('O' + '0' + str(i))

#P10-P15
l93 = ['P' + str(i) for i in range(10, 16)]

l94 = ['P20', 'P21']

l95 = ['P55', 'P56', 'P57.0']

#Q00-Q99
l96 = []
for i in range(100):
    if i > 9:
        l96.append('Q' + str(i))
    else:
        l96.append('Q' + '0' + str(i))

#Q20-Q26
l97 = ['Q' + str(i) for i in range(20, 27)]

#R00-R99
l98 = []
for i in range(100):
    if i > 9:
        l98.append('R' + str(i))
    else:
        l98.append('R' + '0' + str(i))

#S00-T98
l99 = []
for i in range(100):
    if i > 9:
        l99.append('S' + str(i))
        l99.append('T' + str(i))
    else:
        l99.append('S' + '0' + str(i))
        l99.append('T' + '0' + str(i))

'''
WORKING ZONE
'''
def check(code, code_list):
    for i in code_list:
        if i in code:
            return 1
    return 0

def clear_mkb10(code):
    code = code.upper()
    code = code.replace('А', 'A').replace('В', 'B').replace('С', 'C').replace('Д', 'D')
    code = code.replace('Е', 'E').replace('Н', 'H').replace('К', 'K').replace('М', 'M')
    code = code.replace('О', 'O').replace('Р', 'P').replace('Т', 'T').replace('Х', 'X')
    code = code.replace(',', '.').replace('У', 'Y')

    code = code.split()
    if len(code) == 2:
        if len(code[0]) == 3:
            mkb10 = code[1]
        else:
            mkb10 = code[0]
    else:
        mkb10 = code[0]
    return mkb10
    
results = [0 for i in range(100)]
lines = [l2, l3, l4, l5, l6, l7, l8, l9, l10, l11, l12, l13, l14, l15, l16, \
         l17, l18, l19, l20, l21, l22, l23, l24, l25, l26, l27, l28, l29, l30, \
         l31, l32, l33, l34, l35, l36, l37, l38, l39, l40, l41, l42, l43, l44, \
         l45, l46, l47, l48, l49, l50, l51, l52, l53, l54, l55, l56, l57, l58, \
         l59, l60, l61, l62, l63, l64, l65, l66, l67, l68, l69, l70, l71, l72, \
         l73, l74, l75, l76, l77, l78, l79, l80, l81, l82, l83, l84, l85, l86, \
         l87, l88, l89, l90, l91, l92, l93, l94, l95, l96, l97, l98, l99]

res_table = openpyxl.load_workbook('results.xlsx')
res_sheet = res_table['Sheet1']

for patient in patients:
    for i in range(2, 100):
        n = check(clear_mkb10(str(patient.MKB10)), lines[i-2])
        results[i] += n

for i in range(2, 100):
    if res_sheet['C' + str(i)] != None:
        res_sheet['C' + str(i)] = None
    res_sheet['C' + str(i)] = str(results[i])

res_table.save('results.xlsx')

end_time = time.clock()

work_time = round((end_time - start_time), 1)
if work_time < 1:
    work_time = 'менее чем 1'

root = Tk()
root.title('Конец работы')
ttk.Label(root, text='\n\n  Расчет завершен за ' + str(work_time) + ' секунд  \n\n', \
                font=("Arial", 13)).grid()
root.mainloop()